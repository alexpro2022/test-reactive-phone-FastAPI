from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.celery_tasks.tasks import synchronize, task
from app.celery_tasks.utils import (fill_repos, init_repos, is_modified,
                                    read_file)
from app.core import db_flush
from tests import conftest as c
from tests.fixtures import data as d
from tests.utils import compare_lists

FAKE_FILE_PATH = Path('tests/fixtures/Menu.xlsx')


async def service_fills_cache_from_db(service: c.BaseService) -> list:
    assert await service.redis.get_all() is None
    items_db = await service.get_all()  # the method should fill the cache
    assert items_db is not None
    items_cache = await service.redis.get_all()
    assert items_cache is not None
    compare_lists(items_db, items_cache)
    return items_cache


async def _check_repos(get_menu_service: c.MenuService,
                       get_submenu_service: c.SubmenuService,
                       get_dish_service: c.DishService) -> None:

    menus = await service_fills_cache_from_db(get_menu_service)
    assert len(menus) == 2
    for menu in menus:
        assert menu.submenus_count == 2
        assert menu.dishes_count == 6

    submenus = await service_fills_cache_from_db(get_submenu_service)
    assert len(submenus) == 4
    for submenu in submenus:
        assert submenu.dishes_count == 3

    dishes = await service_fills_cache_from_db(get_dish_service)
    assert len(dishes) == 12


def write_file(fname: str, edit: bool = False) -> None:
    wb = load_workbook(filename=fname)
    ws = wb['Лист1']
    if edit:
        ws['B1'] = 'Menu'
    else:
        ws['B1'] = 'Меню'
    wb.save(filename=fname)


def test_menu_file_exists() -> None:
    assert c.FILE_PATH.exists(), f'No such file: {c.FILE_PATH}'


def test_read_file() -> None:
    menus, _, _ = read_file(FAKE_FILE_PATH)
    assert menus == d.EXPECTED_MENU_FILE_CONTENT


@c.pytest_mark_anyio
async def test_db_flush(dish: c.Response,
                        get_menu_repo: c.MenuRepository,
                        get_submenu_repo: c.SubmenuRepository,
                        get_dish_repo: c.DishRepository) -> None:
    assert await get_menu_repo.get_all() is not None
    assert await get_submenu_repo.get_all() is not None
    assert await get_dish_repo.get_all() is not None
    await db_flush(c.engine)
    assert await get_menu_repo.get_all() is None
    # below is redundance but kept just in case
    assert await get_submenu_repo.get_all() is None
    assert await get_dish_repo.get_all() is None


@c.pytest_mark_anyio
async def test_fill_repos(get_menu_service: c.MenuService,
                          get_submenu_service: c.SubmenuService,
                          get_dish_service: c.DishService) -> None:
    assert await get_menu_service.db.get_all() is None
    assert await get_menu_service.redis.get_all() is None
    assert await get_menu_service.get_all() is None
    menus, _, _ = read_file(FAKE_FILE_PATH)
    await fill_repos(menus, get_menu_service, get_submenu_service, get_dish_service)
    await _check_repos(get_menu_service, get_submenu_service, get_dish_service)


@c.pytest_mark_anyio
async def test_init_repos(dish: c.Response,
                          get_test_session: c.AsyncSession,
                          get_test_redis: c.FakeRedis,
                          get_menu_service: c.MenuService,
                          get_submenu_service: c.SubmenuService,
                          get_dish_service: c.DishService) -> None:
    menus = await init_repos(get_test_session, FAKE_FILE_PATH, c.engine, get_test_redis)
    assert menus == d.EXPECTED_MENU_FILE_CONTENT
    await _check_repos(get_menu_service, get_submenu_service, get_dish_service)


@c.pytest_mark_anyio
async def test_task(get_test_session, get_test_redis) -> str:
    msg = 'Меню не изменялось. Выход из фоновой задачи...'
    assert await task(get_test_session, FAKE_FILE_PATH, c.engine) == msg
    write_file(FAKE_FILE_PATH)
    assert await task(get_test_session, FAKE_FILE_PATH, c.engine, get_test_redis) == d.EXPECTED_MENU_FILE_CONTENT


def test_task_name():
    assert synchronize.name == 'app.celery_tasks.tasks.synchronize'


'''
def test_is_modified() -> None:
    assert not is_modified(FAKE_FILE_PATH)
    write_file(FAKE_FILE_PATH)
    assert is_modified(FAKE_FILE_PATH)
'''
